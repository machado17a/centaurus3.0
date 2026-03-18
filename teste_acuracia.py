"""
Plataforma de Teste de Acurácia - Centaurus 3.0
================================================
Testa pares de imagens no sistema de verificação facial e exporta
os resultados para uma planilha Excel/CSV.

Modos de entrada:
  1. Arquivo CSV/Excel com colunas: doc, webcam [, esperado]
  2. Pasta com subpastas: cada subpasta contém exatamente 2 imagens
     (1ª alfabeticamente = documento, 2ª = webcam)

Coluna opcional 'esperado':
  Valores aceitos: 'sim', 'verificado', '1', 'true'  → positivo
                   qualquer outro valor               → negativo
  Quando fornecida, a planilha inclui colunas de Acerto e métricas
  (Acurácia, FAR, FRR, EER estimado).
"""

import csv
import json
import hashlib
import os
import sys
import threading
import traceback
import multiprocessing
from concurrent.futures import ProcessPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Optional

import cv2
import numpy as np
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import filedialog, messagebox
import tkinter as tk

# ── Caminhos do projeto ────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR))

from config.settings import AppConfig
from core.models_loader import ModelsLoader
from core.face_verifier import FaceVerifier


# ── Constantes ─────────────────────────────────────────────────────────────────
POSITIVE_LABELS = {"sim", "s", "verificado", "1", "true", "yes", "y", "positivo"}
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp", ".tiff", ".tif"}

CHECKPOINT_DIR  = BASE_DIR          # Onde salvar os checkpoints
PARTIAL_EVERY   = 500               # Exportação parcial a cada N pares
DEFAULT_WORKERS = max(1, (multiprocessing.cpu_count() or 2) - 1)


# ══════════════════════════════════════════════════════════════════════════════
# Workers — importados do módulo separado (sem tkinter, compatível com spawn)
# ══════════════════════════════════════════════════════════════════════════════
from acuracia_worker import worker_init, worker_task


# ══════════════════════════════════════════════════════════════════════════════
# Checkpoint helpers
# ══════════════════════════════════════════════════════════════════════════════

def _checkpoint_path(folder: str) -> Path:
    """Gera caminho do arquivo de checkpoint baseado na pasta de pares."""
    h = hashlib.md5(folder.encode()).hexdigest()[:10]
    return CHECKPOINT_DIR / f"checkpoint_{h}.json"


def _save_checkpoint(folder: str, results: list[dict]):
    """Salva resultados parciais em JSON."""
    path = _checkpoint_path(folder)
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"folder": folder, "results": results}, f, ensure_ascii=False)


def _load_checkpoint(folder: str) -> list[dict]:
    """Carrega checkpoint se existir, senão retorna lista vazia."""
    path = _checkpoint_path(folder)
    if not path.exists():
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if data.get("folder") == folder:
            return data.get("results", [])
    except Exception:
        pass
    return []


def _delete_checkpoint(folder: str):
    path = _checkpoint_path(folder)
    if path.exists():
        path.unlink()


# ══════════════════════════════════════════════════════════════════════════════
# Lógica de processamento
# ══════════════════════════════════════════════════════════════════════════════

def load_pairs_from_csv(path: Path) -> list[dict]:
    """Lê pares de imagens de um CSV ou Excel."""
    pairs = []

    if path.suffix.lower() in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise ValueError("Planilha vazia.")
            headers = [str(h).strip().lower() if h else "" for h in rows[0]]
            for row in rows[1:]:
                d = dict(zip(headers, row))
                pairs.append({
                    "doc":      str(d.get("doc", d.get("documento", ""))).strip(),
                    "webcam":   str(d.get("webcam", d.get("live", ""))).strip(),
                    "esperado": str(d.get("esperado", d.get("label", ""))).strip(),
                })
        except ImportError:
            raise ImportError("Instale openpyxl para ler arquivos .xlsx: pip install openpyxl")
    else:
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            for row in reader:
                keys = {k.strip().lower(): v for k, v in row.items()}
                pairs.append({
                    "doc":      str(keys.get("doc", keys.get("documento", ""))).strip(),
                    "webcam":   str(keys.get("webcam", keys.get("live", ""))).strip(),
                    "esperado": str(keys.get("esperado", keys.get("label", ""))).strip(),
                })

    return [p for p in pairs if p["doc"] and p["webcam"]]


def load_pairs_from_folder(folder: Path, progress_cb=None) -> list[dict]:
    """
    Descobre pares dentro de uma pasta.
    Estratégia:
      1. Subpastas com exatamente 2 imagens → (imgs[0], imgs[1])
      2. Arquivos planos com sufixo _A/_doc e _B/_webcam (mesmo prefixo)
      3. Fallback: agrupa 2 a 2 por ordem alfabética
    progress_cb(done, total) chamado durante a varredura.
    """
    pairs = []

    # Estratégia 1: subpastas
    subdirs = [d for d in sorted(folder.iterdir()) if d.is_dir()]
    if subdirs:
        total = len(subdirs)
        for i, sub in enumerate(subdirs):
            imgs = sorted([f for f in sub.iterdir() if f.suffix.lower() in IMAGE_EXTS])
            if len(imgs) >= 2:
                pairs.append({"doc": str(imgs[0]), "webcam": str(imgs[1]), "esperado": ""})
            if progress_cb and i % 50 == 0:
                progress_cb(i + 1, total)
        if pairs:
            if progress_cb: progress_cb(total, total)
            return pairs

    # Estratégia 2: dicionário O(n)
    SUFFIX_PAIRS = [
        ("_doc", "_cam"),
        ("_doc", "_webcam"),
        ("_a",   "_b"),
        ("_1",   "_2"),
        ("_ref", "_live"),
    ]

    # Fase 1: contagem rápida para total da barra
    total_files = sum(1 for _ in folder.iterdir())
    if progress_cb: progress_cb(0, total_files)

    # Fase 2: monta índice stem→arquivo
    stem_map: dict[str, Path] = {}
    for i, f in enumerate(folder.iterdir()):
        if f.is_file() and f.suffix.lower() in IMAGE_EXTS:
            stem_map[f.stem.lower()] = f
        if progress_cb and i % 500 == 0:
            progress_cb(i, total_files)
    if progress_cb: progress_cb(total_files, total_files)

    used: set[str] = set()
    for stem_lower, f in sorted(stem_map.items()):
        if stem_lower in used:
            continue
        for suf_a, suf_b in SUFFIX_PAIRS:
            if stem_lower.endswith(suf_a):
                prefix_lower = stem_lower[: -len(suf_a)]
                partner_key  = prefix_lower + suf_b
                partner = stem_map.get(partner_key)
                if partner:
                    pairs.append({"doc": str(f), "webcam": str(partner), "esperado": ""})
                    used.add(stem_lower)
                    used.add(partner_key)
                    break

    if pairs:
        return pairs

    # Estratégia 3: fallback 2 a 2 por ordem alfabética
    all_imgs = sorted(stem_map.values(), key=lambda f: f.name)
    remaining = [f for f in all_imgs if f.stem.lower() not in used]
    for i in range(0, len(remaining) - 1, 2):
        pairs.append({"doc": str(remaining[i]), "webcam": str(remaining[i + 1]), "esperado": ""})

    return pairs


def process_pair(verifier: FaceVerifier, doc_path: str, webcam_path: str) -> dict:
    """Processa um par de imagens e retorna métricas."""
    result = {
        "erro": "",
        "faces_doc": 0,
        "faces_webcam": 0,
        "similaridade_pct": None,
        "cosseno": None,
        "status": "",
    }
    try:
        img_doc = cv2.imread(doc_path)
        img_web = cv2.imread(webcam_path)

        if img_doc is None:
            raise FileNotFoundError(f"Não foi possível abrir: {doc_path}")
        if img_web is None:
            raise FileNotFoundError(f"Não foi possível abrir: {webcam_path}")

        face_doc, emb_doc = verifier.capture_face(img_doc)
        face_web, emb_web = verifier.capture_face(img_web)

        result["faces_doc"] = verifier.count_faces(img_doc)
        result["faces_webcam"] = verifier.count_faces(img_web)

        if emb_doc is None:
            raise ValueError("Nenhuma face detectada no documento")
        if emb_web is None:
            raise ValueError("Nenhuma face detectada na webcam/live")

        vr = verifier.verify_faces(emb_doc, emb_web)
        cos = np.dot(emb_doc, emb_web)
        cos = float(np.clip(cos, -1.0, 1.0))

        result["similaridade_pct"] = round(float(vr.similarity), 4)
        result["cosseno"] = round(cos, 6)
        result["status"] = vr.status

    except Exception as e:
        result["erro"] = str(e)
        result["status"] = "Erro"

    return result


def is_positive(label: str) -> Optional[bool]:
    """Interpreta rótulo esperado; None se vazio."""
    if not label:
        return None
    return label.strip().lower() in POSITIVE_LABELS


def export_results(results: list[dict], out_path: Path):
    """Exporta resultados para Excel (se openpyxl disponível) ou CSV."""
    use_excel = out_path.suffix.lower() == ".xlsx"

    if use_excel:
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Resultados"

            has_labels = any(r.get("esperado") for r in results)

            # ── Cabeçalho ──────────────────────────────────────────────────────
            headers = [
                "#", "Imagem Doc", "Imagem Webcam",
                "Faces Doc", "Faces Webcam",
                "Similaridade (%)", "Cosseno", "Status",
            ]
            if has_labels:
                headers += ["Esperado", "Resultado Esperado", "Acerto"]

            header_fill = PatternFill("solid", fgColor="1a1a2e")
            header_font = Font(bold=True, color="FFFFFF")
            thin = Side(style="thin", color="444444")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            ws.row_dimensions[1].height = 20

            # ── Preenchimento de cores ─────────────────────────────────────────
            COLOR_MAP = {
                "Verificado":               "d4edda",
                "Atenção durante a captura": "fff3cd",
                "Chamar Policial Federal":   "f8d7da",
                "Erro":                      "e2e3e5",
            }
            GREEN  = PatternFill("solid", fgColor="c3e6cb")
            RED    = PatternFill("solid", fgColor="f5c6cb")
            YELLOW = PatternFill("solid", fgColor="ffeeba")
            GRAY   = PatternFill("solid", fgColor="d6d8db")

            # ── Linhas ────────────────────────────────────────────────────────
            tp = tn = fp = fn = 0

            for i, r in enumerate(results, 1):
                row_data = [
                    i,
                    r["doc"],
                    r["webcam"],
                    r["faces_doc"],
                    r["faces_webcam"],
                    r["similaridade_pct"] if r["similaridade_pct"] is not None else "N/A",
                    r["cosseno"] if r["cosseno"] is not None else "N/A",
                    r["status"],
                ]

                acerto_val = ""
                esp_str = ""
                res_esp_str = ""
                if has_labels:
                    esperado_bool = is_positive(r.get("esperado", ""))
                    esp_str = "Verificado" if esperado_bool else ("Não Verificado" if esperado_bool is not None else "")
                    resultado_bool = r["status"] == "Verificado"
                    if esperado_bool is not None:
                        acerto = (resultado_bool == esperado_bool)
                        if esperado_bool and resultado_bool:     tp += 1
                        elif not esperado_bool and not resultado_bool: tn += 1
                        elif not esperado_bool and resultado_bool:     fp += 1
                        else:                                         fn += 1
                        acerto_val = "✓" if acerto else "✗"
                        res_esp_str = "Verificado" if resultado_bool else "Não Verificado"
                    row_data += [esp_str, res_esp_str, acerto_val]

                fill_color = COLOR_MAP.get(r["status"], "FFFFFF")
                row_fill = PatternFill("solid", fgColor=fill_color)

                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=i + 1, column=col, value=val)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    # Colorir coluna status
                    if col == 8:
                        cell.fill = row_fill
                    # Colorir coluna acerto
                    if has_labels and col == len(headers):
                        if acerto_val == "✓":
                            cell.fill = GREEN
                            cell.font = Font(color="155724", bold=True)
                        elif acerto_val == "✗":
                            cell.fill = RED
                            cell.font = Font(color="721c24", bold=True)

            # ── Largura das colunas ────────────────────────────────────────────
            col_widths = [5, 45, 45, 10, 12, 16, 12, 28]
            if has_labels:
                col_widths += [18, 18, 8]
            for col, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = w

            ws.freeze_panes = "A2"

            # ── Aba de Métricas (se há labels) ────────────────────────────────
            if has_labels and (tp + tn + fp + fn) > 0:
                total = tp + tn + fp + fn
                acc    = (tp + tn) / total * 100
                far    = (fp / (fp + tn) * 100) if (fp + tn) > 0 else 0
                frr    = (fn / (fn + tp) * 100) if (fn + tp) > 0 else 0
                eer    = (far + frr) / 2  # estimativa linear simples

                ws2 = wb.create_sheet("Métricas")
                ws2.column_dimensions["A"].width = 35
                ws2.column_dimensions["B"].width = 20

                metrics = [
                    ("Total de pares testados", total),
                    ("Verdadeiros Positivos (TP)", tp),
                    ("Verdadeiros Negativos (TN)", tn),
                    ("Falsos Positivos (FP — FAR)", fp),
                    ("Falsos Negativos (FN — FRR)", fn),
                    ("", ""),
                    ("Acurácia", f"{acc:.2f}%"),
                    ("FAR (False Accept Rate)", f"{far:.2f}%"),
                    ("FRR (False Reject Rate)", f"{frr:.2f}%"),
                    ("EER estimado ((FAR+FRR)/2)", f"{eer:.2f}%"),
                    ("", ""),
                    ("Threshold Verificado (>)", f"{AppConfig.SIMILARITY_THRESHOLD_VERIFIED}%"),
                    ("Threshold Atenção (≥)",   f"{AppConfig.SIMILARITY_THRESHOLD_WARNING}%"),
                ]
                for row_i, (k, v) in enumerate(metrics, 1):
                    ws2.cell(row=row_i, column=1, value=k).font = Font(bold=bool(k))
                    ws2.cell(row=row_i, column=2, value=v)

            wb.save(out_path)
            return

        except ImportError:
            # Fallback para CSV
            out_path = out_path.with_suffix(".csv")

    # ── CSV ──────────────────────────────────────────────────────────────────
    has_labels = any(r.get("esperado") for r in results)
    fieldnames = [
        "#", "doc", "webcam", "faces_doc", "faces_webcam",
        "similaridade_pct", "cosseno", "status",
    ]
    if has_labels:
        fieldnames += ["esperado", "resultado", "acerto"]

    with open(out_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for i, r in enumerate(results, 1):
            row = {
                "#": i,
                "doc": r["doc"],
                "webcam": r["webcam"],
                "faces_doc": r["faces_doc"],
                "faces_webcam": r["faces_webcam"],
                "similaridade_pct": r["similaridade_pct"] if r["similaridade_pct"] is not None else "",
                "cosseno": r["cosseno"] if r["cosseno"] is not None else "",
                "status": r["status"],
            }
            if has_labels:
                esperado_bool = is_positive(r.get("esperado", ""))
                resultado_bool = r["status"] == "Verificado"
                row["esperado"] = "Verificado" if esperado_bool else ("Não Verificado" if esperado_bool is not None else "")
                row["resultado"] = "Verificado" if resultado_bool else "Não Verificado"
                row["acerto"] = ("Sim" if resultado_bool == esperado_bool else "Não") if esperado_bool is not None else ""
            writer.writerow(row)


# ══════════════════════════════════════════════════════════════════════════════
# Interface gráfica
# ══════════════════════════════════════════════════════════════════════════════

class TesteAcuraciaApp(ttk.Window):
    def __init__(self):
        super().__init__(themename="darkly")
        self.title("Centaurus 3.0 — Plataforma de Teste de Acurácia")
        self.geometry("1080x760")
        self.minsize(900, 600)
        self.resizable(True, True)

        self._verifier: Optional[FaceVerifier] = None
        self._pairs: list[dict] = []
        self._results: list[dict] = []
        self._running = False
        self._checkpoint_data: list[dict] = []
        self._pairs_remaining: list[dict] = []
        self._partial_counter: int = 0

        self._build_ui()
        self._load_model_async()

    # ── Construção da UI ──────────────────────────────────────────────────────

    def _build_ui(self):
        # ── Barra de status do modelo ──────────────────────────────────────
        top = ttk.Frame(self, padding=8)
        top.pack(fill=X)

        ttk.Label(top, text="Centaurus 3.0 — Plataforma de Teste de Acurácia",
                  font=("Segoe UI", 13, "bold")).pack(side=LEFT)

        self._model_lbl = ttk.Label(top, text="⏳ Carregando modelo...",
                                    bootstyle="secondary", font=("Segoe UI", 10))
        self._model_lbl.pack(side=RIGHT)

        ttk.Separator(self).pack(fill=X, pady=2)

        # ── Painel de entrada ─────────────────────────────────────────────
        inp = ttk.LabelFrame(self, text="  Entrada  ", padding=10)
        inp.pack(fill=X, padx=12, pady=(6, 0))

        # Linha 1: CSV / Excel
        r1 = ttk.Frame(inp)
        r1.pack(fill=X, pady=2)
        ttk.Label(r1, text="Arquivo CSV/Excel:", width=22).pack(side=LEFT)
        self._csv_var = tk.StringVar()
        ttk.Entry(r1, textvariable=self._csv_var, width=60).pack(side=LEFT, padx=4)
        ttk.Button(r1, text="Procurar…", bootstyle="secondary-outline",
                   command=self._browse_csv).pack(side=LEFT, padx=2)
        ttk.Button(r1, text="Carregar", bootstyle="info-outline",
                   command=self._load_csv).pack(side=LEFT, padx=2)

        ttk.Label(inp, text="  —  OU  —", foreground="gray").pack(anchor=W, pady=2)

        # Linha 2: Pasta
        r2 = ttk.Frame(inp)
        r2.pack(fill=X, pady=2)
        ttk.Label(r2, text="Pasta com subpastas:", width=22).pack(side=LEFT)
        self._folder_var = tk.StringVar()
        ttk.Entry(r2, textvariable=self._folder_var, width=60).pack(side=LEFT, padx=4)
        ttk.Button(r2, text="Procurar…", bootstyle="secondary-outline",
                   command=self._browse_folder).pack(side=LEFT, padx=2)
        ttk.Button(r2, text="Carregar", bootstyle="info-outline",
                   command=self._load_folder).pack(side=LEFT, padx=2)

        # Opções
        opt_frame = ttk.Frame(inp)
        opt_frame.pack(anchor=W, pady=(6, 0))

        self._genuinos_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(
            opt_frame,
            text="Todos os pares são genuínos (mesma pessoa — esperado = Verificado)",
            variable=self._genuinos_var,
            bootstyle="success-round-toggle",
        ).pack(side=LEFT, padx=(0, 16))

        # Info pares + barra de carregamento
        self._pairs_lbl = ttk.Label(inp, text="Nenhum par carregado.",
                                     foreground="gray", font=("Segoe UI", 9))
        self._pairs_lbl.pack(anchor=W, pady=(4, 0))

        load_prog_frame = ttk.Frame(inp)
        load_prog_frame.pack(fill=X, pady=(2, 0))
        self._load_prog_var = tk.DoubleVar(value=0)
        self._load_progress = ttk.Progressbar(
            load_prog_frame, variable=self._load_prog_var,
            maximum=100, bootstyle="info-striped", length=200)
        self._load_prog_lbl = ttk.Label(
            load_prog_frame, text="", width=20, font=("Segoe UI", 8),
            foreground="gray")
        # Oculto por padrão — aparece só durante o carregamento
        self._load_prog_frame = load_prog_frame
        self._load_prog_frame.pack_forget()

        # ── Controles ─────────────────────────────────────────────────────
        ctrl = ttk.Frame(self, padding=(12, 4))
        ctrl.pack(fill=X)

        self._run_btn = ttk.Button(ctrl, text="▶  Iniciar Processamento",
                                   bootstyle="success", width=26,
                                   command=self._start_processing,
                                   state=DISABLED)
        self._run_btn.pack(side=LEFT, padx=(0, 8))

        self._export_btn = ttk.Button(ctrl, text="💾  Exportar Planilha",
                                      bootstyle="primary-outline", width=22,
                                      command=self._export,
                                      state=DISABLED)
        self._export_btn.pack(side=LEFT, padx=4)

        self._stop_btn = ttk.Button(ctrl, text="⏹  Parar",
                                    bootstyle="danger-outline", width=12,
                                    command=self._stop,
                                    state=DISABLED)
        self._stop_btn.pack(side=LEFT, padx=4)

        # ── Opções de processamento ────────────────────────────────────────
        opt2 = ttk.Frame(self, padding=(12, 2))
        opt2.pack(fill=X)

        ttk.Label(opt2, text=f"Workers (CPU × {multiprocessing.cpu_count()} núcleos):",
                  font=("Segoe UI", 9)).pack(side=LEFT)
        self._workers_var = tk.IntVar(value=DEFAULT_WORKERS)
        self._workers_spin = ttk.Spinbox(
            opt2, from_=1, to=multiprocessing.cpu_count(),
            textvariable=self._workers_var, width=4,
            font=("Segoe UI", 9))
        self._workers_spin.pack(side=LEFT, padx=(4, 16))

        ttk.Label(opt2, text="Exportar parcial a cada:",
                  font=("Segoe UI", 9)).pack(side=LEFT)
        self._partial_var = tk.IntVar(value=PARTIAL_EVERY)
        self._partial_spin = ttk.Spinbox(
            opt2, from_=100, to=5000, increment=100,
            textvariable=self._partial_var, width=6,
            font=("Segoe UI", 9))
        self._partial_spin.pack(side=LEFT, padx=(4, 4))
        ttk.Label(opt2, text="pares", font=("Segoe UI", 9)).pack(side=LEFT)

        # Banner de retomada (oculto por padrão)
        self._resume_frame = ttk.Frame(self, padding=(12, 2))
        self._resume_lbl   = ttk.Label(
            self._resume_frame, text="", bootstyle="warning",
            font=("Segoe UI", 9, "bold"))
        self._resume_lbl.pack(side=LEFT)
        self._resume_btn = ttk.Button(
            self._resume_frame, text="↩  Retomar", bootstyle="warning-outline",
            command=self._resume_processing, width=14)
        self._resume_btn.pack(side=LEFT, padx=8)
        self._discard_btn = ttk.Button(
            self._resume_frame, text="🗑  Descartar", bootstyle="danger-outline",
            command=self._discard_checkpoint, width=14)
        self._discard_btn.pack(side=LEFT)

        # ── Barra de progresso ────────────────────────────────────────────
        prog_frame = ttk.Frame(self, padding=(12, 0))
        prog_frame.pack(fill=X)

        self._prog_var = tk.DoubleVar(value=0)
        self._progress = ttk.Progressbar(prog_frame, variable=self._prog_var,
                                          maximum=100, bootstyle="success-striped",
                                          length=400)
        self._progress.pack(side=LEFT, fill=X, expand=True, padx=(0, 8))

        self._prog_lbl = ttk.Label(prog_frame, text="", width=30)
        self._prog_lbl.pack(side=LEFT)

        # ── Tabela de resultados ──────────────────────────────────────────
        tbl_frame = ttk.LabelFrame(self, text="  Resultados  ", padding=6)
        tbl_frame.pack(fill=BOTH, expand=True, padx=12, pady=(4, 0))

        cols = ("#", "Doc", "Webcam", "Sim%", "Cosseno", "Status", "Faces Doc", "Faces Web", "Acerto")
        self._tree = ttk.Treeview(tbl_frame, columns=cols, show="headings",
                                   selectmode="browse")

        widths = [40, 220, 220, 70, 80, 190, 75, 75, 60]
        for col, w in zip(cols, widths):
            self._tree.heading(col, text=col)
            self._tree.column(col, width=w, anchor="center", stretch=(col in ("Doc", "Webcam", "Status")))

        # Tags de cor
        self._tree.tag_configure("ok",   background="#1a3a1a", foreground="#aaffaa")
        self._tree.tag_configure("warn", background="#3a3000", foreground="#ffff88")
        self._tree.tag_configure("fail", background="#3a0000", foreground="#ffaaaa")
        self._tree.tag_configure("err",  background="#2a2a2a", foreground="#aaaaaa")

        vsb = ttk.Scrollbar(tbl_frame, orient=VERTICAL, command=self._tree.yview)
        hsb = ttk.Scrollbar(tbl_frame, orient=HORIZONTAL, command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side=RIGHT, fill=Y)
        hsb.pack(side=BOTTOM, fill=X)
        self._tree.pack(fill=BOTH, expand=True)

        # ── Rodapé de métricas ────────────────────────────────────────────
        self._metrics_lbl = ttk.Label(self, text="", font=("Segoe UI", 9),
                                       foreground="gray", padding=(12, 4))
        self._metrics_lbl.pack(anchor=W)

    # ── Carregamento do modelo ────────────────────────────────────────────────

    def _load_model_async(self):
        def _load():
            try:
                loader = ModelsLoader(AppConfig.get_models_dir())
                model = loader.load_model(AppConfig.MODEL_NAME)
                self._verifier = FaceVerifier(model, AppConfig)
                self.after(0, lambda: self._model_lbl.configure(
                    text="✅ Modelo carregado", bootstyle="success"))
                self.after(0, self._check_can_run)
            except Exception as e:
                self.after(0, lambda: self._model_lbl.configure(
                    text=f"❌ Erro no modelo: {e}", bootstyle="danger"))
        threading.Thread(target=_load, daemon=True).start()

    # ── Ações de entrada ─────────────────────────────────────────────────────

    def _browse_csv(self):
        p = filedialog.askopenfilename(
            title="Selecionar CSV ou Excel",
            filetypes=[("CSV / Excel", "*.csv *.xlsx *.xls"), ("Todos", "*.*")]
        )
        if p:
            self._csv_var.set(p)

    def _browse_folder(self):
        p = filedialog.askdirectory(title="Selecionar pasta com pares de imagens")
        if p:
            self._folder_var.set(p)

    def _load_csv(self):
        path = Path(self._csv_var.get().strip())
        if not path.exists():
            messagebox.showerror("Erro", f"Arquivo não encontrado:\n{path}")
            return
        self._set_loading(True)
        def _work():
            try:
                pairs = load_pairs_from_csv(path)
                self.after(0, lambda: self._finish_load(pairs))
            except Exception as e:
                self.after(0, lambda: self._finish_load(None, str(e)))
        threading.Thread(target=_work, daemon=True).start()

    def _load_folder(self):
        path = Path(self._folder_var.get().strip())
        if not path.is_dir():
            messagebox.showerror("Erro", f"Pasta não encontrada:\n{path}")
            return
        self._set_loading(True)
        def _progress(done, total):
            if total > 0:
                pct = done / total * 100
                self.after(0, lambda p=pct, d=done, t=total: (
                    self._load_prog_var.set(p),
                    self._load_prog_lbl.configure(
                        text=f"{d:,} / {t:,}  ({p:.0f}%)"),
                ))
        def _work():
            try:
                pairs = load_pairs_from_folder(path, progress_cb=_progress)
                self.after(0, lambda: self._finish_load(pairs))
            except Exception as e:
                self.after(0, lambda: self._finish_load(None, str(e)))
        threading.Thread(target=_work, daemon=True).start()

    def _set_loading(self, loading: bool):
        """Ativa/desativa indicador de carregamento."""
        if loading:
            self._pairs_lbl.configure(
                text="⏳  Lendo arquivos da pasta, aguarde...", foreground="#ffdd88")
            self._run_btn.configure(state=DISABLED)
            self._load_prog_var.set(0)
            self._load_prog_lbl.configure(text="")
            self._load_progress.pack(side=LEFT)
            self._load_prog_lbl.pack(side=LEFT, padx=(6, 0))
            self._load_prog_frame.pack(fill=X, pady=(2, 0))
        else:
            self._load_prog_frame.pack_forget()

    def _finish_load(self, pairs, error: str = None):
        """Chamado no thread principal após carregamento."""
        self._set_loading(False)
        if error:
            self._pairs_lbl.configure(text="❌  Erro ao carregar.", foreground="#ff6666")
            messagebox.showerror("Erro ao carregar", error)
            return
        self._pairs = pairs or []
        self._on_pairs_loaded()

    def _on_pairs_loaded(self):
        n = len(self._pairs)
        # Numera cada par com _idx para rastreamento de checkpoint
        for i, p in enumerate(self._pairs):
            p["_idx"] = i + 1
        # Aplica label de genuíno se checkbox marcado e par não tem label
        if self._genuinos_var.get():
            for p in self._pairs:
                if not p.get("esperado"):
                    p["esperado"] = "sim"
        has_labels = any(p.get("esperado") for p in self._pairs)
        lbl = f"✅  {n} par{'es' if n != 1 else ''} carregado{'s' if n != 1 else ''}"
        if has_labels:
            lbl += "  (com rótulos 'esperado' — métricas serão calculadas)"
        self._pairs_lbl.configure(text=lbl, foreground="#88ff88" if n > 0 else "gray")
        self._check_can_run()
        # Limpa tabela
        for item in self._tree.get_children():
            self._tree.delete(item)
        self._results.clear()
        self._export_btn.configure(state=DISABLED)
        self._metrics_lbl.configure(text="")
        # Verifica checkpoint
        self._check_resume()

    def _current_folder(self) -> str:
        """Retorna a pasta atual carregada (ou string do CSV)."""
        return self._folder_var.get().strip() or self._csv_var.get().strip()

    def _check_resume(self):
        """Verifica se há checkpoint para a pasta atual e exibe banner."""
        folder = self._current_folder()
        if not folder:
            self._resume_frame.pack_forget()
            return
        prev = _load_checkpoint(folder)
        if prev:
            done  = len(prev)
            total = len(self._pairs)
            erros = sum(1 for r in prev if r.get("status") == "Erro")
            erro_txt = f"  ⚠ {erros} com erro — recomenda-se Descartar." if erros else ""
            self._resume_lbl.configure(
                text=f"⚠  Checkpoint: {done}/{total} pares já processados.{erro_txt} Deseja retomar?")
            self._resume_frame.pack(fill=X, padx=12, pady=2)
            self._checkpoint_data = prev
        else:
            self._resume_frame.pack_forget()
            self._checkpoint_data = []

    def _resume_processing(self):
        """Retoma processamento do checkpoint."""
        self._resume_frame.pack_forget()
        prev = self._checkpoint_data
        done_idxs = {r["idx"] for r in prev}
        remaining = [p for p in self._pairs if p["_idx"] not in done_idxs]
        self._results = list(prev)
        # Repovoar tabela com resultados anteriores
        for r in self._results:
            self._update_ui(r, r["idx"], r["idx"] / len(self._pairs) * 100, len(self._pairs))
        self._pairs_remaining = remaining
        self._start_processing(resume=True)

    def _discard_checkpoint(self):
        """Descarta checkpoint e começa do zero."""
        folder = self._current_folder()
        _delete_checkpoint(folder)
        self._resume_frame.pack_forget()
        self._checkpoint_data = []

    def _check_can_run(self):
        ready = self._verifier is not None and len(self._pairs) > 0 and not self._running
        self._run_btn.configure(state=NORMAL if ready else DISABLED)

    # ── Processamento paralelo ────────────────────────────────────────────────

    def _start_processing(self, resume: bool = False):
        if not self._pairs or self._verifier is None:
            return
        self._running = True
        if not resume:
            # Apaga checkpoint anterior para não misturar dados de sessões distintas
            _delete_checkpoint(self._current_folder())
            self._results.clear()
            for item in self._tree.get_children():
                self._tree.delete(item)
            self._pairs_remaining = list(self._pairs)
        self._export_btn.configure(state=DISABLED)
        self._run_btn.configure(state=DISABLED)
        self._stop_btn.configure(state=NORMAL)
        self._prog_var.set(0)
        self._metrics_lbl.configure(text="")
        self._partial_counter = len(self._results)
        threading.Thread(target=self._process_thread, daemon=True).start()

    def _process_thread(self):
        folder   = self._current_folder()
        total    = len(self._pairs)
        n_workers = max(1, min(self._workers_var.get(), multiprocessing.cpu_count()))
        partial_n = max(100, self._partial_var.get())
        models_dir = str(AppConfig.get_models_dir())
        model_name = AppConfig.MODEL_NAME

        # Prepara tarefas (somente pares ainda não processados)
        start_idx = len(self._results)  # pares já feitos (retomada)
        tasks = [
            (p.get("_idx", i + 1), p["doc"], p["webcam"], p.get("esperado", ""))
            for i, p in enumerate(getattr(self, "_pairs_remaining", self._pairs))
        ]

        partial_dir = Path(folder) if Path(folder).is_dir() else BASE_DIR
        partial_dir.mkdir(parents=True, exist_ok=True)
        ts_start = datetime.now().strftime("%Y%m%d_%H%M%S")

        with ProcessPoolExecutor(
            max_workers=n_workers,
            initializer=worker_init,
            initargs=(models_dir, model_name),
        ) as executor:
            futures = {executor.submit(worker_task, t): t for t in tasks}
            done_count = start_idx

            for fut in as_completed(futures):
                if not self._running:
                    executor.shutdown(wait=False, cancel_futures=True)
                    break
                try:
                    result = fut.result()
                except Exception as e:
                    t = futures[fut]
                    result = {
                        "idx": t[0], "doc": t[1], "webcam": t[2],
                        "esperado": t[3], "erro": str(e),
                        "faces_doc": 0, "faces_webcam": 0,
                        "similaridade_pct": None, "cosseno": None,
                        "status": "Erro",
                    }

                self._results.append(result)
                done_count += 1
                pct = done_count / total * 100
                self.after(0, lambda r=result, idx=done_count, p=pct:
                           self._update_ui(r, idx, p, total))

                # Checkpoint a cada par
                self._partial_counter += 1
                if self._partial_counter % 50 == 0:
                    _save_checkpoint(folder, self._results)

                # Exportação parcial
                if self._partial_counter % partial_n == 0:
                    p_path = partial_dir / f"parcial_{self._partial_counter:06d}_{ts_start}.xlsx"
                    try:
                        export_results(self._results, p_path)
                    except Exception:
                        pass

        # Checkpoint final
        _save_checkpoint(folder, self._results)
        self.after(0, self._finish_processing)

    def _update_ui(self, r: dict, idx: int, pct: float, total: int):
        self._prog_var.set(pct)
        self._prog_lbl.configure(text=f"{idx}/{total}  ({pct:.0f}%)")

        sim_str = f"{r['similaridade_pct']:.1f}%" if r["similaridade_pct"] is not None else "N/A"
        cos_str = f"{r['cosseno']:.4f}" if r["cosseno"] is not None else "N/A"

        # Mostra erro real na coluna status se houver
        status_display = r["status"]
        if r["status"] == "Erro" and r.get("erro"):
            # Trunca para caber na célula
            status_display = f"Erro: {r['erro'][:60]}"

        esperado_bool = is_positive(r.get("esperado", ""))
        resultado_bool = r["status"] == "Verificado"
        acerto_str = ""
        if esperado_bool is not None:
            acerto_str = "✓" if resultado_bool == esperado_bool else "✗"

        tag = {
            "Verificado": "ok",
            "Atenção durante a captura": "warn",
            "Chamar Policial Federal": "fail",
        }.get(r["status"], "err")

        doc_name = Path(r["doc"]).name
        web_name = Path(r["webcam"]).name

        self._tree.insert("", END, values=(
            idx, doc_name, web_name,
            sim_str, cos_str, status_display,
            r["faces_doc"], r["faces_webcam"],
            acerto_str,
        ), tags=(tag,))
        self._tree.yview_moveto(1.0)

    def _stop(self):
        self._running = False

    def _finish_processing(self):
        self._running = False
        self._stop_btn.configure(state=DISABLED)
        self._check_can_run()

        n = len(self._results)
        erros = sum(1 for r in self._results if r["status"] == "Erro")
        ok    = sum(1 for r in self._results if r["status"] == "Verificado")
        warn  = sum(1 for r in self._results if r["status"] == "Atenção durante a captura")
        fail  = sum(1 for r in self._results if r["status"] == "Chamar Policial Federal")

        # Métricas adicionais se há labels
        has_labels = any(r.get("esperado") for r in self._results)
        metrics_txt = f"Processados: {n}   ✓ {ok}   ⚠ {warn}   ✗ {fail}   Erros: {erros}"
        if has_labels:
            tp = tn = fp = fn = 0
            for r in self._results:
                eb = is_positive(r.get("esperado", ""))
                rb = r["status"] == "Verificado"
                if eb is None: continue
                if eb and rb:      tp += 1
                elif not eb and not rb: tn += 1
                elif not eb and rb:     fp += 1
                else:                   fn += 1
            total_lbl = tp + tn + fp + fn
            if total_lbl > 0:
                acc = (tp + tn) / total_lbl * 100
                far = fp / (fp + tn) * 100 if (fp + tn) > 0 else 0
                frr = fn / (fn + tp) * 100 if (fn + tp) > 0 else 0
                eer = (far + frr) / 2
                metrics_txt += (
                    f"   |   Acurácia: {acc:.1f}%   "
                    f"FAR: {far:.1f}%   FRR: {frr:.1f}%   EER≈{eer:.1f}%"
                )

        self._metrics_lbl.configure(text=metrics_txt)
        self._prog_lbl.configure(text="Concluído ✅")
        self._show_stats_popup()

        if self._results:
            self._export_btn.configure(state=NORMAL)

    # ── Popup de estatísticas ─────────────────────────────────────────────────

    def _show_stats_popup(self):
        if not self._results:
            return

        n = len(self._results)
        erros   = [r for r in self._results if r["status"] == "Erro"]
        validos = [r for r in self._results if r["similaridade_pct"] is not None]

        if not validos:
            return

        sims = [r["similaridade_pct"] for r in validos]
        coss = [r["cosseno"] for r in validos]

        import statistics
        media     = statistics.mean(sims)
        mediana   = statistics.median(sims)
        desvio    = statistics.stdev(sims) if len(sims) > 1 else 0
        minv      = min(sims)
        maxv      = max(sims)
        p10       = sorted(sims)[int(len(sims) * 0.10)]
        p25       = sorted(sims)[int(len(sims) * 0.25)]
        p75       = sorted(sims)[int(len(sims) * 0.75)]
        p90       = sorted(sims)[int(len(sims) * 0.90)]

        ok   = sum(1 for r in self._results if r["status"] == "Verificado")
        warn = sum(1 for r in self._results if r["status"] == "Atenção durante a captura")
        fail = sum(1 for r in self._results if r["status"] == "Chamar Policial Federal")

        # Métricas com labels
        has_labels = any(r.get("esperado") for r in self._results)
        tp = tn = fp = fn = 0
        if has_labels:
            for r in self._results:
                eb = is_positive(r.get("esperado", ""))
                rb = r["status"] == "Verificado"
                if eb is None: continue
                if eb and rb:           tp += 1
                elif not eb and not rb: tn += 1
                elif not eb and rb:     fp += 1
                else:                   fn += 1

        # Janela popup
        pop = tk.Toplevel(self)
        pop.title("Análise Estatística Completa")
        pop.geometry("620x680")
        pop.resizable(False, False)
        pop.grab_set()

        frm = ttk.Frame(pop, padding=16)
        frm.pack(fill=BOTH, expand=True)

        def section(title):
            ttk.Separator(frm).pack(fill=X, pady=(8, 2))
            ttk.Label(frm, text=title, font=("Segoe UI", 10, "bold"),
                      bootstyle="info").pack(anchor=W)

        def row(label, value, color=None):
            f = ttk.Frame(frm)
            f.pack(fill=X, pady=1)
            ttk.Label(f, text=label, width=38, anchor=W,
                      font=("Segoe UI", 9)).pack(side=LEFT)
            lbl = ttk.Label(f, text=str(value), font=("Segoe UI", 9, "bold"),
                            anchor=W)
            if color:
                lbl.configure(foreground=color)
            lbl.pack(side=LEFT)

        ttk.Label(frm, text="Análise Estatística — Pares Agosto Centaurus",
                  font=("Segoe UI", 12, "bold")).pack(anchor=W, pady=(0, 4))

        section("📊 Visão Geral")
        row("Total de pares processados",  n)
        row("Pares com face detectada",     len(validos))
        row("Pares com erro (sem face)",    len(erros),
            "#ff6666" if erros else "#aaffaa")

        section("📈 Distribuição de Similaridade (%)")
        row("Média",         f"{media:.2f}%")
        row("Mediana",       f"{mediana:.2f}%")
        row("Desvio padrão", f"{desvio:.2f}%")
        row("Mínimo",        f"{minv:.2f}%")
        row("Máximo",        f"{maxv:.2f}%")
        row("Percentil 10",  f"{p10:.2f}%")
        row("Percentil 25",  f"{p25:.2f}%")
        row("Percentil 75",  f"{p75:.2f}%")
        row("Percentil 90",  f"{p90:.2f}%")

        section("🎯 Decisões pelo Sistema")
        t = AppConfig.SIMILARITY_THRESHOLD_VERIFIED
        w = AppConfig.SIMILARITY_THRESHOLD_WARNING
        row(f"✓  Verificado  (sim% > {t})",
            f"{ok}  ({ok/n*100:.1f}%)", "#88ff88")
        row(f"⚠  Atenção     ({w}–{t}%)",
            f"{warn}  ({warn/n*100:.1f}%)", "#ffff88")
        row(f"✗  Alerta       (sim% < {w})",
            f"{fail}  ({fail/n*100:.1f}%)", "#ff8888")
        row("Erros (sem detecção)",
            f"{len(erros)}  ({len(erros)/n*100:.1f}%)", "#aaaaaa")

        if has_labels and (tp + tn + fp + fn) > 0:
            total_lbl = tp + tn + fp + fn
            acc = (tp + tn) / total_lbl * 100
            far = fp / (fp + tn) * 100 if (fp + tn) > 0 else 0
            frr = fn / (fn + tp) * 100 if (fn + tp) > 0 else 0
            eer = (far + frr) / 2
            prec = tp / (tp + fp) * 100 if (tp + fp) > 0 else 0
            rec  = tp / (tp + fn) * 100 if (tp + fn) > 0 else 0
            f1   = 2 * prec * rec / (prec + rec) if (prec + rec) > 0 else 0

            section("🔬 Métricas de Classificação (com rótulos esperados)")
            row("Verdadeiros Positivos (TP)",  tp)
            row("Verdadeiros Negativos (TN)",  tn)
            row("Falsos Positivos (FP)",        fp, "#ff8888" if fp > 0 else None)
            row("Falsos Negativos (FN)",        fn, "#ffaa44" if fn > 0 else None)
            row("Acurácia",  f"{acc:.2f}%",  "#88ff88" if acc >= 80 else "#ffaa44")
            row("Precisão",  f"{prec:.2f}%")
            row("Recall (Sensibilidade)", f"{rec:.2f}%")
            row("F1-Score",  f"{f1:.2f}%")
            row("FAR (False Accept Rate)",  f"{far:.2f}%",
                "#ff8888" if far > 5 else "#88ff88")
            row("FRR (False Reject Rate)",  f"{frr:.2f}%",
                "#ffaa44" if frr > 10 else "#88ff88")
            row("EER estimado ((FAR+FRR)/2)", f"{eer:.2f}%",
                "#ff8888" if eer > 10 else "#ffaa44")

            # Análise de threshold
            section("🔧 Análise de Threshold")
            row(f"Threshold atual Verificado", f"> {t}%")
            row(f"Threshold atual Atenção",    f"≥ {w}%")
            below_70 = sum(1 for s in sims if s <= t)
            below_65 = sum(1 for s in sims if s < w)
            row(f"Pares genuínos com sim% ≤ {t} (subotimais)",
                f"{below_70}  ({below_70/len(validos)*100:.1f}%)",
                "#ffaa44" if below_70 > 0 else "#88ff88")
            row(f"Pares genuínos com sim% < {w} (rejeitados)",
                f"{below_65}  ({below_65/len(validos)*100:.1f}%)",
                "#ff8888" if below_65 > 0 else "#88ff88")

        ttk.Separator(frm).pack(fill=X, pady=(10, 6))
        ttk.Button(frm, text="Fechar", bootstyle="secondary",
                   command=pop.destroy).pack(anchor=E)

    # ── Exportação ────────────────────────────────────────────────────────────

    def _export(self):
        if not self._results:
            return

        try:
            import openpyxl
            has_xl = True
        except ImportError:
            has_xl = False

        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"acuracia_{ts}.xlsx" if has_xl else f"acuracia_{ts}.csv"

        filetypes = []
        if has_xl:
            filetypes.append(("Excel", "*.xlsx"))
        filetypes.append(("CSV", "*.csv"))

        path = filedialog.asksaveasfilename(
            title="Salvar planilha de resultados",
            defaultextension=".xlsx" if has_xl else ".csv",
            initialfile=default_name,
            filetypes=filetypes,
        )
        if not path:
            return

        try:
            export_results(self._results, Path(path))
            messagebox.showinfo("Exportado", f"Planilha salva em:\n{path}")
        except Exception as e:
            messagebox.showerror("Erro ao exportar", f"{e}\n\n{traceback.format_exc()}")


# ══════════════════════════════════════════════════════════════════════════════
# Entry point
# ══════════════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    multiprocessing.freeze_support()  # necessário para spawn no Windows
    app = TesteAcuraciaApp()
    app.mainloop()
