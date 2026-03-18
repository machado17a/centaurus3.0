"""
Separador de Pares por Status — Centaurus 3.0
==============================================
Lê um checkpoint JSON (ou planilha Excel/CSV de resultado) e organiza os pares
em subpastas por status, gerando:

  • Imagem composta  doc | similaridade | webcam  (fácil inspeção visual)
  • Planilha Excel resumo por grupo + resumo geral

Uso:
    python separar_por_status.py                        # auto-detecta checkpoint
    python separar_por_status.py resultado.xlsx         # a partir de planilha
    python separar_por_status.py checkpoint_xxx.json    # checkpoint específico
    python separar_por_status.py --so-excel             # só gera Excel, sem imagens

Saída:  separacao_YYYYMMDD_HHMMSS/
            verde_verificado/
            amarelo_atencao/
            vermelho_pf/
            erro_na/
            resumo_geral.xlsx
"""

import argparse
import json
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import cv2
import numpy as np

# ── Configuração de grupos ─────────────────────────────────────────────────────
GRUPOS = {
    "Verificado":                ("verde_verificado",    ( 46, 166,  40), "✓"),
    "Atenção durante a captura": ("amarelo_atencao",     ( 50, 200, 200), "⚠"),
    "Chamar Policial Federal":   ("vermelho_pf",         ( 41,  41, 207), "✗"),
    "Erro":                      ("erro_na",             (120, 120, 120), "?"),
}
FALLBACK_GRUPO = ("outros",                              (160, 160, 160), "?")

BASE_DIR = Path(__file__).parent

# ── Helpers ────────────────────────────────────────────────────────────────────

def _load_from_checkpoint(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)
    return data.get("results", [])


def _load_from_excel(path: Path) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise ImportError("Instale openpyxl: pip install openpyxl")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h else "" for h in rows[0]]
    results = []
    for row in rows[1:]:
        d = dict(zip(headers, row))
        results.append({
            "idx":             d.get("#", d.get("idx", len(results) + 1)),
            "doc":             str(d.get("imagem doc", d.get("doc", ""))),
            "webcam":          str(d.get("imagem webcam", d.get("webcam", ""))),
            "status":          str(d.get("status", "")),
            "similaridade_pct": d.get("similaridade (%)", d.get("similaridade_pct")),
            "cosseno":         d.get("cosseno"),
            "faces_doc":       d.get("faces doc", d.get("faces_doc", 0)),
            "faces_webcam":    d.get("faces webcam", d.get("faces_webcam", 0)),
            "erro":            str(d.get("erro", "")),
            "esperado":        str(d.get("esperado", "")),
        })
    return results


def _load_from_csv(path: Path) -> list[dict]:
    import csv
    results = []
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            keys = {k.strip().lower(): v for k, v in row.items()}
            results.append({
                "idx":             int(keys.get("#", keys.get("idx", len(results)+1)) or 0),
                "doc":             str(keys.get("imagem doc", keys.get("doc", ""))),
                "webcam":          str(keys.get("imagem webcam", keys.get("webcam", ""))),
                "status":          str(keys.get("status", "")),
                "similaridade_pct": keys.get("similaridade (%)", keys.get("similaridade_pct")),
                "cosseno":         keys.get("cosseno"),
                "faces_doc":       keys.get("faces doc", keys.get("faces_doc", 0)),
                "faces_webcam":    keys.get("faces webcam", keys.get("faces_webcam", 0)),
                "erro":            str(keys.get("erro", "")),
                "esperado":        str(keys.get("esperado", "")),
            })
    return results


def auto_detect_source() -> Path:
    """Escolhe automaticamente: checkpoint mais recente, depois xlsx, depois csv."""
    candidates = sorted(BASE_DIR.glob("checkpoint_*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]
    for pattern in ("resultado*.xlsx", "parcial_*.xlsx", "resultado*.csv"):
        found = sorted(BASE_DIR.glob(pattern), key=lambda p: p.stat().st_mtime, reverse=True)
        if found:
            return found[0]
    raise FileNotFoundError(
        "Nenhum checkpoint ou arquivo de resultado encontrado.\n"
        "Passe o arquivo como argumento: python separar_por_status.py <arquivo>"
    )


def make_composite(img_doc: np.ndarray, img_webcam: np.ndarray,
                   status: str, sim: float | None) -> np.ndarray:
    """Cria imagem composta: doc | painel central | webcam."""
    TARGET_H = 256

    def _resize(img):
        if img is None or img.size == 0:
            return np.full((TARGET_H, 240, 3), 80, dtype=np.uint8)
        h, w = img.shape[:2]
        scale = TARGET_H / max(h, 1)
        return cv2.resize(img, (max(1, int(w * scale)), TARGET_H))

    left  = _resize(img_doc)
    right = _resize(img_webcam)

    # Painel central com status + similaridade
    cfg = GRUPOS.get(status, (None, FALLBACK_GRUPO[1], FALLBACK_GRUPO[2]))
    color_bgr = cfg[1]
    panel_w = 180
    panel = np.full((TARGET_H, panel_w, 3), 30, dtype=np.uint8)

    # Barra colorida no centro
    cv2.rectangle(panel, (10, TARGET_H//2 - 60), (panel_w - 10, TARGET_H//2 + 60), color_bgr, -1)

    sim_txt = f"{sim:.1f}%" if sim is not None else "N/A"

    # Texto similaridade
    cv2.putText(panel, sim_txt,
                (panel_w//2 - 35, TARGET_H//2 - 15),
                cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255, 255, 255), 2, cv2.LINE_AA)

    # Status curto
    short = {"Verificado": "VERIF.", "Atenção durante a captura": "ATENÇÃO",
             "Chamar Policial Federal": "REPROV.", "Erro": "ERRO"}.get(status, status[:8])
    cv2.putText(panel, short,
                (panel_w//2 - len(short)*5, TARGET_H//2 + 28),
                cv2.FONT_HERSHEY_SIMPLEX, 0.48, (255, 255, 255), 1, cv2.LINE_AA)

    # Rótulos doc / webcam
    label_bar_h = 28
    def _add_label(img, label):
        bar = np.full((label_bar_h, img.shape[1], 3), 20, dtype=np.uint8)
        cv2.putText(bar, label, (6, 20), cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200, 200, 200), 1)
        return np.vstack([bar, img])

    left  = _add_label(left,  "DOCUMENTO")
    right = _add_label(right, "WEBCAM / LIVE")
    pad_h = left.shape[0]
    panel_pad = np.full((pad_h, panel_w, 3), 30, dtype=np.uint8)
    panel_pad[label_bar_h:, :] = panel

    # Borda vertical separadora
    sep = np.full((pad_h, 4, 3), 60, dtype=np.uint8)
    return np.hstack([left, sep, panel_pad, sep, right])


def export_group_excel(group_results: list[dict], path: Path, status_name: str):
    """Gera planilha Excel para um grupo de resultados."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = status_name[:31]

    headers = ["#", "Arquivos Doc / Webcam", "Similitude (%)", "Cosseno",
               "Faces Doc", "Faces Webcam", "Esperado", "Erro"]
    hfill = PatternFill("solid", fgColor="1a1a2e")
    hfont = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")

    for i, r in enumerate(group_results, 1):
        pair_txt = f"{Path(r['doc']).name}  /  {Path(r['webcam']).name}"
        ws.append([
            r.get("idx", i),
            pair_txt,
            round(float(r["similaridade_pct"]), 2) if r.get("similaridade_pct") is not None else None,
            round(float(r["cosseno"]), 6)           if r.get("cosseno")          is not None else None,
            r.get("faces_doc", 0),
            r.get("faces_webcam", 0),
            r.get("esperado", ""),
            r.get("erro", ""),
        ])

    # Largura das colunas
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 60
    for col in "CDEFGH":
        ws.column_dimensions[col].width = 18

    wb.save(path)


def export_summary_excel(all_results: list[dict], path: Path, stats: dict):
    """Gera planilha resumo geral com todas as entradas + aba de estatísticas."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        return

    COLOR_MAP = {
        "Verificado":                "d4edda",
        "Atenção durante a captura": "fff3cd",
        "Chamar Policial Federal":   "f8d7da",
        "Erro":                      "e2e3e5",
    }

    wb = openpyxl.Workbook()

    # ── Aba Resumo ──────────────────────────────────────────────────────────────
    ws_stat = wb.active
    ws_stat.title = "Resumo"
    ws_stat.append(["Status", "Quantidade", "% do Total"])
    total = sum(stats.values())
    hfill = PatternFill("solid", fgColor="1a1a2e")
    hfont = Font(bold=True, color="FFFFFF")
    for c in ws_stat[1]:
        c.fill = hfill; c.font = hfont; c.alignment = Alignment(horizontal="center")
    for status, n in sorted(stats.items(), key=lambda x: -x[1]):
        pct = n / total * 100 if total else 0
        ws_stat.append([status, n, round(pct, 2)])
    ws_stat.append(["TOTAL", total, 100.0])
    for col in "ABC":
        ws_stat.column_dimensions[col].width = 35

    # ── Aba Todos os Resultados ─────────────────────────────────────────────────
    ws = wb.create_sheet("Todos os Resultados")
    headers = ["#", "Imagem Doc", "Imagem Webcam", "Status",
               "Similaridade (%)", "Cosseno", "Faces Doc", "Faces Webcam", "Erro"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hfill; c.font = hfont
        c.alignment = Alignment(horizontal="center")

    for r in sorted(all_results, key=lambda x: x.get("idx", 0)):
        fill_hex = COLOR_MAP.get(r.get("status", ""), "FFFFFF")
        row_vals = [
            r.get("idx"),
            Path(r["doc"]).name    if r.get("doc")    else "",
            Path(r["webcam"]).name if r.get("webcam") else "",
            r.get("status", ""),
            round(float(r["similaridade_pct"]), 2) if r.get("similaridade_pct") is not None else None,
            round(float(r["cosseno"]),          6) if r.get("cosseno")          is not None else None,
            r.get("faces_doc", 0),
            r.get("faces_webcam", 0),
            r.get("erro", ""),
        ]
        ws.append(row_vals)
        row_fill = PatternFill("solid", fgColor=fill_hex)
        for cell in ws[ws.max_row]:
            cell.fill = row_fill

    for w, col in zip([8, 38, 38, 28, 18, 14, 12, 12, 50], "ABCDEFGHI"):
        ws.column_dimensions[col].width = w

    wb.save(path)
    print(f"  → Excel geral: {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Separa pares de faces por status")
    parser.add_argument("source", nargs="?", help="Checkpoint JSON, Excel ou CSV de resultado")
    parser.add_argument("--so-excel", action="store_true", help="Não copia imagens, só gera Excel")
    parser.add_argument("--saida", default=None, help="Pasta de saída (padrão: separacao_TIMESTAMP)")
    args = parser.parse_args()

    # ── Carrega dados ──────────────────────────────────────────────────────────
    if args.source:
        src = Path(args.source)
    else:
        src = auto_detect_source()

    print(f"Fonte: {src}")
    suffix = src.suffix.lower()
    if suffix == ".json":
        results = _load_from_checkpoint(src)
    elif suffix in (".xlsx", ".xls"):
        results = _load_from_excel(src)
    elif suffix == ".csv":
        results = _load_from_csv(src)
    else:
        raise ValueError(f"Formato não suportado: {suffix}")

    print(f"Total de resultados: {len(results)}")
    if not results:
        print("Nenhum resultado encontrado. Encerrando.")
        return

    # ── Pasta de saída ─────────────────────────────────────────────────────────
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.saida) if args.saida else BASE_DIR / f"separacao_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)
    print(f"Saída: {out_dir}\n")

    # ── Agrupa por status ──────────────────────────────────────────────────────
    by_status: dict[str, list[dict]] = defaultdict(list)
    for r in results:
        by_status[r.get("status", "")].append(r)

    stats = {s: len(v) for s, v in by_status.items()}
    total = len(results)
    print("Distribuição:")
    for s, n in sorted(stats.items(), key=lambda x: -x[1]):
        pct = n / total * 100
        print(f"  {s or '(vazio)':<32} {n:5d}  ({pct:.1f}%)")
    print()

    # ── Processa cada grupo ────────────────────────────────────────────────────
    for status, group_results in sorted(by_status.items(), key=lambda x: -len(x[1])):
        cfg = GRUPOS.get(status)
        folder_name = cfg[0] if cfg else "outros"
        grp_dir = out_dir / folder_name
        grp_dir.mkdir(exist_ok=True)

        print(f"[{folder_name}]  {len(group_results)} pares")

        # Planilha do grupo
        export_group_excel(group_results, grp_dir / "lista.xlsx", status or "outros")
        print(f"  → lista.xlsx gerada")

        if not args.so_excel:
            import gc
            erros_img = 0
            BATCH = 50   # libera memória a cada N imagens
            for i, r in enumerate(group_results):
                doc_path    = Path(r.get("doc",    ""))
                webcam_path = Path(r.get("webcam", ""))
                idx = r.get("idx", 0)

                stem    = doc_path.stem if doc_path.name else f"{idx:06d}"
                out_img = grp_dir / f"{stem}_comp.jpg"

                if out_img.exists():          # já gerada (retomada)
                    continue

                img_doc    = None
                img_webcam = None
                try:
                    if doc_path.exists():
                        img_doc    = cv2.imread(str(doc_path))
                    if webcam_path.exists():
                        img_webcam = cv2.imread(str(webcam_path))

                    sim = r.get("similaridade_pct")
                    composite = make_composite(img_doc, img_webcam, status,
                                               float(sim) if sim is not None else None)
                    cv2.imwrite(str(out_img), composite,
                                [cv2.IMWRITE_JPEG_QUALITY, 88])
                except Exception:
                    erros_img += 1
                finally:
                    del img_doc, img_webcam
                    if "composite" in dir():
                        del composite

                if (i + 1) % BATCH == 0:
                    gc.collect()
                    print(f"    {i + 1}/{len(group_results)}...", end="\r")

            gc.collect()
            if erros_img:
                print(f"  ⚠  {erros_img} composições falharam (imagens ausentes/corrompidas)")
            print(f"  → {len(group_results) - erros_img} composições geradas em {grp_dir.name}/")

        print()

    # ── Excel geral ────────────────────────────────────────────────────────────
    export_summary_excel(results, out_dir / "resumo_geral.xlsx", stats)
    print(f"\nConcluído! Pasta de saída: {out_dir}")

    # Abre a pasta no Explorer
    import subprocess
    subprocess.Popen(["explorer", str(out_dir)])


if __name__ == "__main__":
    main()
